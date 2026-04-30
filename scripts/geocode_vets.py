import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl, requests, json, time

# Excelを読み込み
wb = openpyxl.load_workbook('data/tokyo23_animalhospital.xlsx')
ws = wb.active

hospitals = []
for r in range(2, ws.max_row + 1):
    name    = ws.cell(r, 1).value
    address = ws.cell(r, 4).value  # 住所列
    if not name or not address:
        continue
    hospitals.append({'name': str(name), 'address': str(address)})

print(f'読み込み件数: {len(hospitals)}件')

# 国土地理院ジオコーダーAPIで緯度経度を取得
# https://msearch.gsi.go.jp/address-search/AddressSearch?q=住所
GEOCODE_URL = 'https://msearch.gsi.go.jp/address-search/AddressSearch'

results = []
failed = []

for i, h in enumerate(hospitals):
    try:
        resp = requests.get(
            GEOCODE_URL,
            params={'q': h['address']},
            timeout=10
        )
        data = resp.json()
        if data:
            coords = data[0]['geometry']['coordinates']
            lng, lat = coords[0], coords[1]
            results.append({
                'name': h['name'],
                'lat': lat,
                'lng': lng,
                'address': h['address'],
            })
        else:
            failed.append(h)
            print(f'  ジオコード失敗: {h["name"]} / {h["address"]}')
    except Exception as e:
        failed.append(h)
        print(f'  エラー: {h["name"]}: {e}')

    if (i + 1) % 100 == 0:
        print(f'  {i+1}/{len(hospitals)}件処理完了')

    time.sleep(0.1)  # レートリミット対策

print(f'\n成功: {len(results)}件 / 失敗: {len(failed)}件')

# data/vets.jsonに保存（上書き）
with open('data/vets.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print(f'data/vets.json 更新完了: {len(results)}件')

# 失敗リストを保存
if failed:
    with open('data/vets_failed.json', 'w', encoding='utf-8') as f:
        json.dump(failed, f, ensure_ascii=False, indent=2)
    print(f'失敗リスト: data/vets_failed.json ({len(failed)}件)')
